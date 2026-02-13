import os
import glob
import pandas as pd
import tkinter as tk
from tkinter import messagebox
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def feldolgozas(ev, honap):
    mappa_utvonal = r"C:\Users\szabo\Desktop\pythonscripts\gyar"
    arlista_fajl = os.path.join(mappa_utvonal, "arlista.xlsx")
    if not os.path.exists(arlista_fajl):
        messagebox.showerror("Hiba", "Nem található arlista.xlsx!")
        return

    arlista = pd.read_excel(arlista_fajl)
    arlista["Dátum"] = pd.to_datetime(arlista["Dátum"]).dt.date

    osszesitett_df = pd.DataFrame()
    fajlok = glob.glob(os.path.join(mappa_utvonal, "*forgalom_telephely*.xls*"))
    if not fajlok:
        messagebox.showwarning("Figyelem", "Nem található megfelelő fájl.")
        return

    for fajl in fajlok:
        try:
            telephely = os.path.splitext(os.path.basename(fajl))[0].split("forgalom_telephely_")[-1]
            df = pd.read_excel(fajl, header=0)
            df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
            if df.empty:
                continue

            datum_oszlop = None
            for col in df.columns:
                if pd.to_datetime(df[col], errors='coerce').notna().any():
                    datum_oszlop = col
                    df[col] = pd.to_datetime(df[col]).dt.date
                    break
            if datum_oszlop is None:
                continue

            menny_oszlop = None
            for col in df.columns:
                if "mennyiség" in str(col).lower():
                    menny_oszlop = col
                    break
            if menny_oszlop is None:
                messagebox.showerror("Hiba", f"A {fajl} fájlban nem található 'Mennyiség' oszlop.")
                return

            if "Termék" not in df.columns:
                messagebox.showerror("Hiba", f"A {fajl} fájlban nincs 'Termék' oszlop.")
                return

            df = df[(pd.to_datetime(df[datum_oszlop]).dt.year == ev) &
                    (pd.to_datetime(df[datum_oszlop]).dt.month == honap)].copy()
            if df.empty:
                continue

            def ar_hozzarend(datum, termek):
                lejar = arlista[arlista["Dátum"] <= datum]
                if lejar.empty or termek not in lejar.columns:
                    return 0
                return lejar[termek].iloc[(lejar["Dátum"] <= datum).sum() - 1]

            df["Ár"] = df.apply(lambda x: ar_hozzarend(x[datum_oszlop], x["Termék"]), axis=1)
            df["Bevétel"] = df[menny_oszlop] * df["Ár"]

            df["Telephely"] = telephely
            oszlopok_rend = [c for c in df.columns if c != "Telephely"] + ["Telephely"]
            df = df[oszlopok_rend]

            osszesitett_df = pd.concat([osszesitett_df, df], ignore_index=True)

        except Exception as e:
            print(f"Hiba a fájl feldolgozásánál: {fajl}\n{e}")
            continue

    if osszesitett_df.empty:
        messagebox.showinfo("Eredmény", "Nincs adat a megadott időszakra.")
        return

    kimeneti_fajlnev = f"forgalomkimutatas_{ev}{honap:02d}.xlsx"
    kimeneti_utvonal = os.path.join(mappa_utvonal, kimeneti_fajlnev)

    if os.path.exists(kimeneti_utvonal):
        answer = messagebox.askyesno("Figyelem", f"A fájl már létezik:\n{kimeneti_utvonal}\nFelülírod?")
        if not answer:
            return

    with pd.ExcelWriter(kimeneti_utvonal, engine="openpyxl") as writer:

        # ===== ÖSSZESÍTŐ =====
        osszesito = osszesitett_df.groupby("Telephely")[[menny_oszlop, "Bevétel"]].sum().reset_index()
        osszesito.to_excel(writer, sheet_name="Összesítő", index=False)

        # ===== NAPI BONTÁS =====
        pivot_nap = osszesitett_df.groupby([datum_oszlop, "Telephely"])[[menny_oszlop, "Bevétel"]].sum().reset_index()
        pivot_nap.to_excel(writer, sheet_name="Napi bontás", index=False)

        # ===== TELEPHELY LAPOK =====
        for telephely in osszesitett_df["Telephely"].unique():
            telephely_df = osszesitett_df[osszesitett_df["Telephely"] == telephely].copy()
            sheet_nev = str(telephely)[:31]
            telephely_df.to_excel(writer, sheet_name=sheet_nev, index=False)

            sheet = writer.book[sheet_nev]
            last_row = sheet.max_row + 1

            # Mindösszesen sor
            sheet[f"A{last_row}"] = "Mindösszesen"

            for idx, cell in enumerate(sheet[1], start=1):
                if cell.value == menny_oszlop:
                    col = get_column_letter(idx)
                    sheet[f"{col}{last_row}"] = f"=SUM({col}2:{col}{last_row-1})"
                if cell.value == "Bevétel":
                    col = get_column_letter(idx)
                    sheet[f"{col}{last_row}"] = f"=SUM({col}2:{col}{last_row-1})"
                    sheet[f"{col}{last_row}"].number_format = '#,##0 "Ft"'

        workbook = writer.book

        # ===== FORMÁZÁS =====
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for cell in sheet[1]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in sheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                sheet.column_dimensions[col_letter].width = max_length + 2

            for idx, cell in enumerate(sheet[1], start=1):
                if cell.value == "Bevétel":
                    col_letter = get_column_letter(idx)
                    for row in range(2, sheet.max_row + 1):
                        sheet[f"{col_letter}{row}"].number_format = '#,##0 "Ft"'

            for idx, cell in enumerate(sheet[1], start=1):
                if cell.value == datum_oszlop:
                    col_letter = get_column_letter(idx)
                    for row in range(2, sheet.max_row + 1):
                        sheet[f"{col_letter}{row}"].number_format = 'YYYY-MM-DD'

        # ===== ÖSSZESÍTŐ MINDÖSSZESEN =====
        summary_sheet = workbook["Összesítő"]
        last_row = summary_sheet.max_row + 1
        summary_sheet[f"A{last_row}"] = "Mindösszesen"
        summary_sheet[f"B{last_row}"] = f"=SUM(B2:B{last_row-1})"
        summary_sheet[f"C{last_row}"] = f"=SUM(C2:C{last_row-1})"
        summary_sheet[f"C{last_row}"].number_format = '#,##0 "Ft"'

        # ===== NAPI BONTÁS MINDÖSSZESEN =====
        pivot_sheet = workbook["Napi bontás"]
        last_row = pivot_sheet.max_row + 1
        pivot_sheet[f"A{last_row}"] = "Mindösszesen"
        pivot_sheet[f"C{last_row}"] = f"=SUM(C2:C{last_row-1})"
        pivot_sheet[f"D{last_row}"] = f"=SUM(D2:D{last_row-1})"
        pivot_sheet[f"D{last_row}"].number_format = '#,##0 "Ft"'

    messagebox.showinfo("Kész", f"Elkészült:\n{kimeneti_utvonal}")


# ===== GUI =====
def inditas():
    try:
        evhonap = evhonap_entry.get().strip()
        if len(evhonap) != 6 or not evhonap.isdigit():
            raise ValueError("Az időszakot YYYYMM formátumban add meg (pl. 202602)")
        ev = int(evhonap[:4])
        honap = int(evhonap[4:])
        if honap < 1 or honap > 12:
            raise ValueError("A hónap 1 és 12 között legyen.")
        ablak.destroy()
        feldolgozas(ev, honap)
    except Exception as e:
        messagebox.showerror("Hiba", str(e))

ablak = tk.Tk()
ablak.title("Forgalom kimutatás")

tk.Label(ablak, text="ÉvHónap (YYYYMM, pl. 202602):").grid(row=0, column=0, padx=10, pady=5)
evhonap_entry = tk.Entry(ablak)
evhonap_entry.grid(row=0, column=1, padx=10, pady=5)
tk.Button(ablak, text="Indítás", command=inditas).grid(row=1, columnspan=2, pady=10)
ablak.mainloop()
